from openpyxl import Workbook, load_workbook

# define data class
class info:
    account = "undefined"
    name = "undefined"
    location = "undefined"
    salesAmount = 0.0
    profitRate = 0.0
    normal = True      # to mark the cell is nornal or not


# open the file
data = load_workbook("A.xlsx",read_only=True, data_only=True)

# for each file to process data 


# to store this file data which platform and type is
platform = "wish"
infoType = "A类"

# use info list to temporary store data 
infoList = []


# for each table to process data
sheets = data._sheets
for table in sheets:
    row = 1
    nrow = table.max_row

    # for each module to find out useful info
    while row < nrow:
        # add 1 to row to ignore the title
        row += 1

        # create the module info instance
        infoInstance = info()

        infoInstance.name = table.cell(row,2).value.split('/',2)[0]
        if infoInstance.name == "":
            infoInstance.name = '/'

        # to process the account/location cell
        row_3 = table.cell(row,3).value
        infoInstance.account = row_3.split('/')[0]
        # remove the charactor "仓"
        infoInstance.location = row_3.split('/')[1][0:-1]       

        # to judge this module is normal or not
        row_4 = table.cell(row,4).value
        if row_4 == None:
            infoInstance.normal = False
            infoInstance.salesAmount = table.cell(row,5).value
        else :
            infoInstance.salesAmount = row_4
            infoInstance.profitRate = table.cell(row+1,5).value

        # append the module instance into the list
        infoList.append(infoInstance)

        # add 3 to variable row to move to next module 
        row += 3




# open report to write down
summary = load_workbook("B.xlsx")

# find the infotype table
reportTable = summary[infoType]

# store the merge cell's info
mergeList = reportTable.merged_cells
mergeDict = {}
for mergeCell in mergeList:
    mergeDict[mergeCell.min_row] = mergeCell.max_row - mergeCell.min_row

# to find the index of platform
index = 1
while index < reportTable.max_row:
    if reportTable["C" + str(index)].value.lower() == platform:
        break
    else :
        if mergeDict.get(index) != None:
            index += mergeDict.get(index) + 1
        else :
            index +=1

# for each data in infoList to write down in the report
for infoInstance in infoList:
    isFind = False      # to mark the account is finded or not

    for row in range(index, index + mergeDict[index]):
        reportAccount = reportTable["D" + str(row)].value
        reportLocationList = reportTable["F" + str(row)].value.split(' ')
        reportLocation = reportLocationList[1] if len(reportLocationList) > 1 else reportLocationList[0]

        # print(infoInstance.account + "\t" + reportAccount + "\t" + infoInstance.location + "\t" + reportLocation)
        # to match corret account and location row
        if infoInstance.account == reportAccount and infoInstance.location == reportLocation :
            isFind = True

            # if the name is wrong, then change the name 
            reportName = reportTable["E" + str(row)].value
            if infoInstance.name !=  reportName : 
                reportTable["E" + str(row)].value = infoInstance.name
                reportTable["F" + str(row)].value = infoInstance.name + " " + reportLocation

            # to judge the normal is true or not
            if infoInstance.normal :
                # if normal, then write down the salesAmount and profitRate
                reportTable["G" + str(row)].value = infoInstance.salesAmount
                reportTable["H" + str(row)].value = infoInstance.profitRate
            else :
                # else write down the margin 
                reportTable["J" + str(row)].value = infoInstance.salesAmount

            # print(row)

            # write down and break the for loop
            break

    if isFind:
        continue

    # 由于有合并表格的存在，插入一行真的极其的烦，功能后面在迭代吧，我不行了
    # if do not find the account in the table, then create it
    # reportTable.insert_rows(index)
    # reportTable["D" + str(index)].value = infoInstance.account
    # reportTable["E" + str(index)].value = infoInstance.name
    # reportTable["F" + str(index)].value = infoInstance.name + " " + infoInstance.location
    #  # to judge the normal is true or not
    # if infoInstance.normal :
    #     # if normal, then write down the salesAmount and profitRate
    #     reportTable["G" + str(index)].value = infoInstance.salesAmount
    #     reportTable["H" + str(index)].value = infoInstance.profitRate
    # else :
    #     # else write down the margin 
    #     reportTable["J" + str(index)].value = infoInstance.salesAmount

    # if do not find the account in the table, then print
    print()

summary.save("output.xlsx")
            


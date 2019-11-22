from error import ErrorList, Error, Warning, processException
import sys
from xlrd import open_workbook
from openpyxl import load_workbook,Workbook
import profit
import saleAmount
import refund
import utils
import notfound


def doCommand():
    # forever loop
    while True:
        # split the command and remove the none
        cmdLine = input("(self)>>>")
        if cmdLine == "" : continue

        cmdList = cmdLine.strip().split(" ")
        for cmdIteration in cmdList:
            if cmdIteration == '' : cmdList.remove(cmdIteration)

        # we have no different cmd now, so ignore it
        # cmd = cmdList[0]

        index = 1
        result = {}
        while index < len(cmdList):
            # if it is the path Argument, then save it
            if cmdList[index] in pathArguDict:
                index += 1
                if index >= len(cmdList):
                    print("data none!!" + cmdList[index-1])
                    break
                else:
                    result[pathArguDict[cmdList[index-1]]] = cmdList[index]
                    index += 1

            elif cmdList[index] in systemArguDict:
                systemArguDict[cmdList[index]]()
                break

            else:
                print("command error! " + cmdList[index])
                break

        if "salePath" in result and "summaryPath" in result and "refundPath" in result:
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process salesAmount
            SA = saleAmount.SaleAmount(summary, notfoundTable)
            SA.processDir(result["salePath"])

            # process refund
            RF = refund.Refund(summary,notfoundTable)
            RF.processRefundInfo(result["refundPath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            summary.save(savePath)
            NF.save()

        elif "salePath" in result and "summaryPath" in result and "refundPath" not in result:
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process salesAmount
            SA = saleAmount.SaleAmount(summary, notfoundTable)
            SA.processDir(result["salePath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            summary.save(savePath)

        elif "summaryPath" in result and "refundPath" in result:
            try:
                summary = load_workbook(result["summaryPath"])
            except:
                processException()

            # get the notfound table
            notfoundPath = result["notfoundTable"] if "notfoundPath" in result else ""
            NF = notfound.NotFound(notfoundPath)
            notfoundTable = NF.getNotfoundTable()

            # process refund
            RF = refund.Refund(summary,notfoundTable)
            RF.processRefundInfo(result["refundPath"])

            # save file
            savePath = result["savePath"] if "savePath" in result else "summary.xlsx"
            summary.save(savePath)
            

        elif "originPath" in result and "updatePath" in result:
            PF = profit.Profit(result["originPath"],result["updatePath"])
            PF.processProfitUpdate()

            savePath = result["savePath"] if "savePath" in result else ""
            PF.save(savePath)

        # print
        if len(result) != 0 : ErrorList.printErrorList()


def printVersion():
    print("6.0 By XiaoMing \n")

def printHelp():
    for pa in pathArguDict.keys():
        print(("%-10.5s" % pa) + pathArguDict[pa])

    print(("%-10.5s" % "-v") + "查看版本")
    print(("%-10.5s" % "-help") + "查看帮助")
    print(("%-10.5s" % "-exit") + "退出程序")

    print("for example: do -o filePath -u filePath")



cmdDict = {
    # "do" : 
}

pathArguDict = {
    "-o" : "originPath", 
    "-u" : "updatePath",
    "-sa" : "salePath",
    "-r" : "refundPath",
    "-su" : "summaryPath",
    "-out" : "savePath",
    "-n" : "notfoundPath"
}

systemArguDict = {
    "-v" : printVersion,
    "-help" : printHelp,
    "-exit" : sys.exit
}


        

    

